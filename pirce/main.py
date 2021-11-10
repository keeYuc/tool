#! /Users/keeyu/homebrew/bin/python3

import json
import requests
import openpyxl
import os
import datetime
import time
import copy


def count_time(func):
    def int_time(*args, **kwargs):
        start_time = datetime.datetime.now()  # 程序开始时间
        func()
        over_time = datetime.datetime.now()   # 程序结束时间
        total_time = (over_time-start_time).total_seconds()
        print('程序共计%s秒' % total_time)
    return int_time


def count_init_timer(class_):
    class wrapper():
        def __init__(self, *args, **kwargs):
            start_time = datetime.datetime.now()  # 程序开始时间
            self.wrapped = class_(*args, **kwargs)
            over_time = datetime.datetime.now()   # 程序结束时间
            total_time = (over_time-start_time).total_seconds()
            print('初始化共计%s秒' % total_time)

        def __getattr__(self, name):
            return getattr(self.wrapped, name)
    return wrapper


@count_init_timer
class bk_updater():
    def __init__(self) -> None:
        self.list = self.__read_code()
        self.map = {}
        self.f_map = {}
        self.__run()

    def __url_make(self, list):
        url = ''
        url_1 = 'http://api.waizaowang.com/doc/getStockBKDayKLine?code='
        url_2 = '&startDate=2021-01-01&endDate=2100-01-01&ktype=103&fields=code,name,ktype,fq,tdate,open,close,high,low,cjl,cje,hsl&export=1&token=febb869f0979d084c4a8d17ce45ea866'
        for i in list:
            if len(url) == 0:
                url = url_1 + i
            else:
                url = url+','+i
        url += url_2
        return url

    def __load_data(self):
        for i in self.list:
            rsb = requests.get(self.__url_make(i))
            for i in json.loads(rsb.text)['data']:
                if i['code'] in self.map.keys():
                    self.map[i['code']].append(i)
                else:
                    self.map[i['code']] = [i]

    def __calc_data(self):
        for k in self.map:
            self.f_map[k] = {'y_open': self.map[k][0]
                             ['open'], 'y_open_item': self.map[k][0], 'm_open': self.map[k][-1]['open'], 'm_item': self.map[k][-1], 'm_close': self.map[k][-1]['close'], 'code': k, 'name': self.map[k][0]
                             ['name'], }

    def __run(self):
        self.__load_data()
        self.__calc_data()
        self.__write_execl()

    def __read_code(self):
        list = []
        tmp = []
        with open('code.json', 'r') as f:
            for i in json.load(f)['data']:
                if len(tmp) < 30:
                    tmp.append(i['code'])
                else:
                    list.append(tmp)
                    tmp = []
        return list

    def __write_execl(self):
        path = '操作记录.xlsx'
        wb = openpyxl.load_workbook(path)
        ws = wb['板块强度']
        start = 7
        for k in self.f_map:
            ws['A{}'.format(start)] = k
            ws['B{}'.format(start)] = self.f_map[k]['name']
            # 年涨幅
            ws['C{}'.format(start)] = 100*(self.f_map[k]['m_close'] -
                                           self.f_map[k]['y_open'])/self.f_map[k]['y_open']
            # 月涨幅
            ws['D{}'.format(start)] = 100*(self.f_map[k]['m_close'] -
                                           self.f_map[k]['m_open'])/self.f_map[k]['m_open']
            # 年较强度
            ws['E{}'.format(start)] = '=(C{})/(E2+100)*100'.format(start)
            # 月较强度
            ws['F{}'.format(start)] = '=(D{})/(F2+100)*100'.format(start)
            # 年排名
            ws['G{}'.format(start)] = '=RANK(E{},E7:E399,0)'.format(start)
            # 月排名
            ws['H{}'.format(start)] = '=RANK(F{},F7:F399,0)'.format(start)

            ws['J{}'.format(start)] = '{}'.format(
                self.f_map[k]['y_open_item']['tdate'])
            ws['K{}'.format(start)] = '{}'.format(
                self.f_map[k]['m_item']['tdate'])

            start += 1
        os.remove(path)
        wb.save(path)


a = bk_updater()
